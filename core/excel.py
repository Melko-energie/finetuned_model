"""Excel exports: single-sheet batch view and per-supplier multi-sheet view."""

import io

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from core.config import FIELDS_LEFT, FIELDS_RIGHT, ALL_FIELD_KEYS


def export_excel_batch(results_list: list[dict]) -> bytes:
    """Single sheet, one row per invoice, color-coded by status."""
    all_labels = ["Nom du PDF", "Type"] + [lbl for _, lbl in FIELDS_LEFT + FIELDS_RIGHT] + ["Fournisseur"]
    rows = []
    for res in results_list:
        row = {"Nom du PDF": res["filename"]}
        row["Type"] = "AVOIR" if res.get("is_avoir") else "FACTURE"
        for key, label in FIELDS_LEFT + FIELDS_RIGHT:
            val = res["fields"].get(key)
            if val is None or val == "null":
                val = ""
            elif isinstance(val, list):
                val = " | ".join(str(v) for v in val)
            row[label] = str(val)
        row["Fournisseur"] = res.get("installateur", "").upper()
        rows.append(row)
    df = pd.DataFrame(rows, columns=all_labels)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extractions", startrow=2)
        ws = writer.sheets["Extractions"]

        nb_cols = len(all_labels)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
        title_cell = ws.cell(row=1, column=1, value=f"Extraction batch - {len(results_list)} factures")
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="1E3A5F")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        for col_idx in range(1, nb_cols + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 32

        val_font = Font(name="Calibri", size=11)
        val_fill_ok = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        val_fill_empty = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
        val_fill_avoir = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        val_fill_warn = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        val_font_warn = Font(name="Calibri", size=11, italic=True, color="E65100")
        label_to_key = {lbl: key for key, lbl in FIELDS_LEFT + FIELDS_RIGHT}

        for row_idx in range(4, 4 + len(results_list)):
            res = results_list[row_idx - 4]
            fields = res.get("fields") or {}
            a_verifier = fields.get("_a_verifier", [])
            is_avoir = res.get("is_avoir", False)

            for col_idx in range(1, nb_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = val_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                col_label = all_labels[col_idx - 1]
                field_key = label_to_key.get(col_label, "")

                if is_avoir:
                    cell.fill = val_fill_avoir
                    if col_idx == 1:
                        cell.font = Font(name="Calibri", size=11, bold=True)
                    elif not cell.value:
                        cell.value = "Non detecte"
                        cell.font = Font(name="Calibri", size=11, italic=True, color="C53030")
                elif col_idx == 1:
                    cell.font = Font(name="Calibri", size=11, bold=True)
                elif field_key in a_verifier:
                    cell.fill = val_fill_warn
                    cell.font = val_font_warn
                    if cell.value:
                        cell.value = str(cell.value) + " (a verifier)"
                elif not cell.value:
                    cell.fill = val_fill_empty
                    cell.value = "Non detecte"
                    cell.font = Font(name="Calibri", size=11, italic=True, color="C53030")
                else:
                    cell.fill = val_fill_ok
            ws.row_dimensions[row_idx].height = 26

        for col_idx, label in enumerate(all_labels, 1):
            max_len = len(label)
            for row_idx in range(4, 4 + len(results_list)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = max_len + 4

    return buf.getvalue()


def export_excel_multi_sheets(results_list: list[dict]) -> bytes:
    """One sheet per supplier + a TOUTES_FACTURES summary sheet."""
    columns = [
        "NOM_FICHIER", "TYPE", "NUMERO_FACTURE", "DATE_FACTURE",
        "MONTANT_HT", "TAUX_TVA", "MONTANT_TTC",
        "NOM_INSTALLATEUR", "COMMUNE_TRAVAUX", "CODE_POSTAL", "ADRESSE_TRAVAUX",
    ]
    field_keys = ALL_FIELD_KEYS

    def build_rows(items):
        rows = []
        for res in items:
            row = {"NOM_FICHIER": res["filename"]}
            row["TYPE"] = "AVOIR" if res.get("is_avoir") else "FACTURE"
            for col, key in zip(columns[2:], field_keys):
                val = res["fields"].get(key)
                if val is None or val == "null":
                    val = ""
                elif isinstance(val, list):
                    val = " | ".join(str(v) for v in val)
                row[col] = str(val)
            rows.append(row)
        return rows

    by_inst = {}
    for res in results_list:
        inst = res.get("installateur", "INCONNU").upper()
        by_inst.setdefault(inst, []).append(res)

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    val_font = Font(name="Calibri", size=11)
    err_font = Font(name="Calibri", size=11, italic=True, color="C53030")
    err_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
    ok_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    warn_font = Font(name="Calibri", size=11, italic=True, color="E65100")
    avoir_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    col_to_key = dict(zip(columns[2:], field_keys))

    def style_sheet(ws, rows_data, items):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=ws.title)
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="1E3A5F")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 32

        for row_idx in range(4, 4 + len(rows_data)):
            res = items[row_idx - 4]
            fields = res.get("fields") or {}
            a_verifier = fields.get("_a_verifier", [])
            is_avoir = res.get("is_avoir", False)

            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                col_name = columns[col_idx - 1]
                field_key = col_to_key.get(col_name, "")

                if is_avoir:
                    cell.fill = avoir_fill
                    if col_idx == 1:
                        cell.font = Font(name="Calibri", size=11, bold=True)
                    elif not cell.value:
                        cell.value = "Non detecte"
                        cell.font = err_font
                    else:
                        cell.font = val_font
                elif col_idx == 1:
                    cell.font = Font(name="Calibri", size=11, bold=True)
                elif field_key in a_verifier:
                    cell.fill = warn_fill
                    cell.font = warn_font
                    if cell.value:
                        cell.value = str(cell.value) + " (a verifier)"
                elif not cell.value:
                    cell.value = "Non detecte"
                    cell.font = err_font
                    cell.fill = err_fill
                else:
                    cell.font = val_font
                    cell.fill = ok_fill
            ws.row_dimensions[row_idx].height = 26

        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(col_name)
            for row_idx in range(4, 4 + len(rows_data)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = max_len + 4

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        all_rows = build_rows(results_list)
        df_all = pd.DataFrame(all_rows, columns=columns)
        df_all.to_excel(writer, index=False, sheet_name="TOUTES_FACTURES", startrow=2)
        style_sheet(writer.sheets["TOUTES_FACTURES"], all_rows, results_list)

        for inst_name in sorted(by_inst.keys()):
            sheet_name = inst_name[:31]
            inst_items = by_inst[inst_name]
            inst_rows = build_rows(inst_items)
            df_inst = pd.DataFrame(inst_rows, columns=columns)
            df_inst.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
            style_sheet(writer.sheets[sheet_name], inst_rows, inst_items)

    return buf.getvalue()
