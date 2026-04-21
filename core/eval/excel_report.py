"""Render an evaluation result as a colored XLSX with two sheets:
Summary (metrics overview) and Details (one row per PDF, color-coded)."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.config import ALL_FIELD_KEYS

# Palette — aligned with core/excel.py conventions
_FILL_MATCH = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
_FILL_MISMATCH = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
_FILL_MISSING = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
_FILL_UNEXPECTED = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
_FILL_META = PatternFill(start_color="F2F4F6", end_color="F2F4F6", fill_type="solid")
_FILL_HEADER = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

_FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
_FONT_TITLE = Font(name="Calibri", size=14, bold=True, color="1E3A5F")
_FONT_VAL = Font(name="Calibri", size=11)
_FONT_ERR = Font(name="Calibri", size=11, italic=True, color="C53030")

_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

_VERDICT_FILL = {
    "match": _FILL_MATCH,
    "mismatch": _FILL_MISMATCH,
    "missing": _FILL_MISSING,
    "unexpected": _FILL_UNEXPECTED,
}


def dump_excel(result: dict, output_path: Path) -> None:
    """Write the full evaluation result to an XLSX with Summary / Details
    / Per-Supplier sheets."""
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet
    _build_summary(wb, result)
    _build_details(wb, result)
    if result.get("metrics_by_supplier"):
        _build_per_supplier(wb, result)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


# ─── Per-Supplier sheet (chantier 2.4) ────────────────────────

def _acc_fill(acc: float) -> PatternFill:
    """Traffic-light palette for accuracy cells."""
    if acc >= 0.9:
        return _FILL_MATCH      # green
    if acc >= 0.6:
        return _FILL_MISSING    # orange
    return _FILL_MISMATCH       # red


def _build_per_supplier(wb: Workbook, result: dict) -> None:
    by_supplier = result["metrics_by_supplier"]
    ws = wb.create_sheet("Per-Supplier")

    headers = ["Fournisseur", "N PDFs", "Accuracy micro", "Accuracy macro"] + list(ALL_FIELD_KEYS)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "B2"

    # Sort by micro accuracy ASC (worst first)
    sorted_items = sorted(
        by_supplier.items(),
        key=lambda kv: kv[1]["global"]["accuracy"],
    )

    for r, (supplier, m) in enumerate(sorted_items, start=2):
        gg = m["global"]
        per_field = m["per_field"]

        # Meta cells (no color)
        cell = ws.cell(row=r, column=1, value=supplier)
        cell.fill = _FILL_META
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="left", vertical="center")

        cell = ws.cell(row=r, column=2, value=m["n_pdfs"])
        cell.fill = _FILL_META
        cell.font = _FONT_VAL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")

        # Global micro accuracy (colored by threshold)
        micro = gg["accuracy"]
        cell = ws.cell(row=r, column=3, value=f"{micro * 100:.1f}%")
        cell.fill = _acc_fill(micro)
        cell.font = _FONT_VAL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")

        # Global macro accuracy (colored by threshold)
        macro = gg["accuracy_macro"]
        cell = ws.cell(row=r, column=4, value=f"{macro * 100:.1f}%")
        cell.fill = _acc_fill(macro)
        cell.font = _FONT_VAL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal="center")

        # Per-field accuracies (colored)
        for i, field in enumerate(ALL_FIELD_KEYS, start=5):
            acc = per_field[field]["accuracy"]
            cell = ws.cell(row=r, column=i, value=f"{acc * 100:.1f}%")
            cell.fill = _acc_fill(acc)
            cell.font = _FONT_VAL
            cell.border = _BORDER
            cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    for i in range(5, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20


# ─── Summary sheet ──────────────────────────────────────────

def _build_summary(wb: Workbook, result: dict) -> None:
    ws = wb.create_sheet("Summary")
    meta = result["meta"]
    per_field = result["metrics"]["per_field"]
    g = result["metrics"]["global"]

    ws.merge_cells("A1:G1")
    title = ws.cell(row=1, column=1, value="Evaluation Report")
    title.font = _FONT_TITLE
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Metadata block (rows 3-9)
    rows_meta = [
        ("Dataset", meta["pdfs_dir"]),
        ("Ground truth", meta["truth_file"]),
        ("Matched PDFs", meta["matched"]),
        ("Missing on disk", meta["missing_on_disk"]),
        ("Skipped (not in truth)", meta["skipped_no_truth"]),
        ("Model", meta["model"]),
        ("Started at", meta["started_at"]),
        ("Duration (s)", round(meta["duration_seconds"], 1)),
    ]
    for i, (k, v) in enumerate(rows_meta, start=3):
        kcell = ws.cell(row=i, column=1, value=k)
        kcell.font = Font(name="Calibri", size=10, bold=True, color="434654")
        ws.cell(row=i, column=2, value=v).font = _FONT_VAL

    # Per-field metrics block
    header_row = 3 + len(rows_meta) + 2
    headers = ["Field", "Match", "Mismatch", "Missing", "Unexpected", "Total", "Accuracy"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center")
        cell.border = _BORDER
    ws.row_dimensions[header_row].height = 28

    for idx, field in enumerate(ALL_FIELD_KEYS, start=header_row + 1):
        c = per_field[field]
        values = [field, c["match"], c["mismatch"], c["missing"], c["unexpected"],
                  c["total"], f"{c['accuracy'] * 100:.1f}%"]
        for col, v in enumerate(values, start=1):
            cell = ws.cell(row=idx, column=col, value=v)
            cell.border = _BORDER
            cell.font = _FONT_VAL
            cell.alignment = Alignment(horizontal="center" if col > 1 else "left")

    # Global accuracy block
    global_row = header_row + len(ALL_FIELD_KEYS) + 2
    ws.cell(row=global_row, column=1, value="Global micro accuracy").font = Font(bold=True)
    ws.cell(row=global_row, column=2, value=f"{g['accuracy'] * 100:.1f}%").font = _FONT_VAL
    ws.cell(row=global_row, column=3, value=f"({g['match']} / {g['total']} cells)").font = _FONT_VAL
    ws.cell(row=global_row + 1, column=1, value="Global macro accuracy").font = Font(bold=True)
    ws.cell(row=global_row + 1, column=2, value=f"{g['accuracy_macro'] * 100:.1f}%").font = _FONT_VAL
    ws.cell(row=global_row + 1, column=3, value="(mean of per-field)").font = _FONT_VAL

    # Column widths
    widths = [26, 14, 14, 14, 14, 10, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─── Details sheet ──────────────────────────────────────────

def _build_details(wb: Workbook, result: dict) -> None:
    ws = wb.create_sheet("Details")

    meta_headers = ["Nom du PDF", "Type", "Installateur détecté", "Erreur"]
    field_headers = []
    for field in ALL_FIELD_KEYS:
        field_headers.extend([f"{field}\n(expected)", f"{field}\n(extracted)"])
    headers = meta_headers + field_headers

    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "E2"

    for r, row in enumerate(result["per_pdf"], start=2):
        expected = row["expected"] or {}
        extracted = row["extracted"] or {}
        verdicts = row["verdicts"] or {}
        err = row.get("error")

        # Metadata columns (gray neutral)
        meta_vals = [row["filename"], _expected_type(expected, row), row.get("installateur") or "", err or ""]
        for c, v in enumerate(meta_vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = _FILL_META
            cell.font = _FONT_ERR if c == 4 and v else _FONT_VAL
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = _BORDER

        # Field columns (color per verdict)
        col = len(meta_headers) + 1
        for field in ALL_FIELD_KEYS:
            verdict = verdicts.get(field, "mismatch")
            fill = _VERDICT_FILL.get(verdict, _FILL_MISMATCH)
            for value in (expected.get(field), extracted.get(field)):
                cell = ws.cell(row=r, column=col, value=_safe_str(value))
                cell.fill = fill
                cell.font = _FONT_VAL
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.border = _BORDER
                col += 1

    # Column widths
    ws.column_dimensions["A"].width = 34  # filename
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 30
    for i in range(len(meta_headers) + 1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22


def _expected_type(expected: dict, row: dict) -> str:
    t = row.get("type") or ""
    if t:
        return t
    if expected.get("Type"):
        return expected["Type"]
    return ""


def _safe_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, list):
        return " | ".join(str(x) for x in v)
    return str(v)
