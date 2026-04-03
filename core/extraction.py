"""Extraction logic — no Streamlit dependencies."""

import io
import os
import json
import zipfile
import importlib.util
from pathlib import Path

import numpy as np
from PIL import Image
import fitz
import pandas as pd
import ollama

PROJECT_ROOT = Path(__file__).resolve().parent.parent

# Import from scripts/12_gemma2_smart.py
_spec_smart = importlib.util.spec_from_file_location(
    "gemma2_smart", str(PROJECT_ROOT / "scripts" / "12_gemma2_smart.py")
)
_mod_smart = importlib.util.module_from_spec(_spec_smart)
_spec_smart.loader.exec_module(_mod_smart)

detect_installateur = _mod_smart.detect_installateur
detect_avoir = _mod_smart.detect_avoir
extraire_champs = _mod_smart.extraire_champs
inverser_montants_avoir = _mod_smart.inverser_montants_avoir
get_ocr_text = _mod_smart.get_ocr_text
PROMPTS_INSTALLATEURS = _mod_smart.PROMPTS_INSTALLATEURS

# OCR model singleton
_ocr_model = None


def get_ocr_model():
    global _ocr_model
    if _ocr_model is None:
        from doctr.models import ocr_predictor
        _ocr_model = ocr_predictor(
            det_arch="db_resnet50", reco_arch="crnn_vgg16_bn", pretrained=True
        )
    return _ocr_model


# ─────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────

FIELDS_LEFT = [
    ("NUMERO_FACTURE", "Numero Facture"),
    ("DATE_FACTURE", "Date Facture"),
    ("MONTANT_HT", "Montant HT"),
    ("TAUX_TVA", "Taux TVA"),
    ("MONTANT_TTC", "Montant TTC"),
]
FIELDS_RIGHT = [
    ("NOM_INSTALLATEUR", "Installateur"),
    ("COMMUNE_TRAVAUX", "Commune"),
    ("CODE_POSTAL", "Code Postal"),
    ("ADRESSE_TRAVAUX", "Adresse Travaux"),
]
ALL_FIELD_KEYS = [k for k, _ in FIELDS_LEFT + FIELDS_RIGHT]

PROMPT_TEXTE = """Tu es un expert comptable specialise dans les factures BTP francaises.

REGLES :
- NUMERO_FACTURE : numero COMPLET apres "N", "Facture N", "Ref". Garde tous les caracteres.
- DATE_FACTURE : format JJ/MM/AAAA.
- MONTANT_HT : montant AVANT TVA. Format "358.83". Cherche "Total HT", "Net HT".
- TAUX_TVA : uniquement le pourcentage. Exemple "20%".
- MONTANT_TTC : montant FINAL avec TVA. Cherche "Total TTC", "Net a payer".
- NOM_INSTALLATEUR : nom COMPLET de l'entreprise emettrice (pas le client).
- COMMUNE_TRAVAUX : ville du chantier.
- CODE_POSTAL : code postal a 5 chiffres.
- ADRESSE_TRAVAUX : adresse complete du chantier.

TEXTE DE LA FACTURE :
{texte}

Reponds UNIQUEMENT en JSON valide. Aucun texte avant ou apres.
{{"NUMERO_FACTURE":null,"DATE_FACTURE":null,"MONTANT_HT":null,"TAUX_TVA":null,"MONTANT_TTC":null,"NOM_INSTALLATEUR":null,"COMMUNE_TRAVAUX":null,"CODE_POSTAL":null,"ADRESSE_TRAVAUX":null}}"""


def clean_json(raw):
    raw = raw.strip()
    if "```json" in raw:
        raw = raw.split("```json")[1].split("```")[0].strip()
    elif "```" in raw:
        raw = raw.split("```")[1].split("```")[0].strip()
    return raw


# ─────────────────────────────────────────
# EXTRACTION FUNCTIONS
# ─────────────────────────────────────────

def extract_from_precomputed_ocr(filename: str) -> dict:
    """Extract using pre-computed OCR text + generic Gemma2 prompt.
    Returns {"fields": dict, "error": str|None, "is_avoir": bool}
    """
    ocr_dir = PROJECT_ROOT / "data" / "ocr_texts"
    base_name = os.path.splitext(filename)[0]
    base_name = base_name.replace("_page0", "").replace("_page_001", "")

    texte = ""
    for root, dirs, files in os.walk(str(ocr_dir)):
        for f in files:
            if base_name.lower() in f.lower():
                try:
                    with open(os.path.join(root, f), "r", encoding="utf-8") as fp:
                        data = json.load(fp)
                        pages_ocr = data.get("pages", [])
                        for page in pages_ocr:
                            if isinstance(page, list):
                                lignes = {}
                                for token in page:
                                    if isinstance(token, dict):
                                        text = token.get("text", "").strip()
                                        if not text:
                                            continue
                                        bbox = token.get("bbox", [0, 0, 0, 0])
                                        y = round(bbox[1] / 10) * 10
                                        if y not in lignes:
                                            lignes[y] = []
                                        lignes[y].append((bbox[0], text))
                                for y in sorted(lignes.keys()):
                                    tokens_ligne = sorted(lignes[y], key=lambda t: t[0])
                                    texte += "  ".join(t[1] for t in tokens_ligne) + "\n"
                except Exception:
                    pass
                break

    if not texte:
        return {"fields": None, "error": "OCR introuvable pour cette facture", "is_avoir": False}

    is_avoir = detect_avoir(texte)

    prompt = PROMPT_TEXTE.format(texte=texte[:3000])
    response = ollama.chat(
        model="gemma2:9b",
        options={"temperature": 0, "seed": 42},
        messages=[{"role": "user", "content": prompt}],
    )

    try:
        result = json.loads(clean_json(response["message"]["content"]))
        if is_avoir and result:
            result = inverser_montants_avoir(result)
        return {"fields": result, "error": None, "is_avoir": is_avoir}
    except json.JSONDecodeError:
        return {"fields": None, "error": "Reponse JSON invalide", "is_avoir": False}


def extract_smart(filename: str, fournisseur: str = "Auto-detect") -> dict:
    """Extract using pre-computed OCR + smart supplier-specific prompt.
    Returns {"fields": dict, "error": str|None, "installateur": str, "is_avoir": bool}
    """
    texte = get_ocr_text(filename)
    if texte is None:
        return {
            "fields": None,
            "error": "OCR introuvable pour cette facture",
            "installateur": None,
            "is_avoir": False,
        }

    if fournisseur == "Auto-detect":
        installateur = detect_installateur(texte)
    elif fournisseur == "DEFAULT":
        installateur = "DEFAULT"
    else:
        installateur = fournisseur.lower()

    is_avoir = detect_avoir(texte)
    fields = extraire_champs(texte, installateur)
    if is_avoir and fields:
        fields = inverser_montants_avoir(fields)

    return {
        "fields": fields,
        "error": None,
        "installateur": installateur,
        "is_avoir": is_avoir,
    }


def process_file_live(file_bytes: bytes, suffix: str, fournisseur: str = "Auto-detect") -> dict:
    """Full pipeline: DocTR OCR in memory + Gemma2 extraction.
    Returns {"fields": dict, "error": str|None, "installateur": str, "is_avoir": bool}
    """
    pil_images = []
    if suffix == ".pdf":
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        for page in doc:
            pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            pil_images.append(img)
        doc.close()
    else:
        pil_images.append(Image.open(io.BytesIO(file_bytes)).convert("RGB"))

    nb_pages = len(pil_images)

    model_ocr = get_ocr_model()
    np_images = [np.array(img) for img in pil_images]
    result_ocr = model_ocr(np_images)

    MONTANT_KEYWORDS = ["ht", "tva", "ttc", "total", "net", "payer", "montant"]
    if nb_pages == 1:
        pages_indices = [0]
    else:
        best_page = nb_pages - 1
        best_score = 0
        for pi in range(1, nb_pages):
            words = [
                word.value.lower()
                for block in result_ocr.pages[pi].blocks
                for line in block.lines
                for word in line.words
            ]
            score = sum(1 for w in words for kw in MONTANT_KEYWORDS if kw in w)
            if score > best_score or (score == best_score and pi > best_page):
                best_score = score
                best_page = pi
        pages_indices = [0] if best_page == 0 else [0, best_page]

    all_tokens = []
    for pi in pages_indices:
        page = result_ocr.pages[pi]
        h, w = pil_images[pi].height, pil_images[pi].width
        for block in page.blocks:
            for line in block.lines:
                for word in line.words:
                    x0, y0 = word.geometry[0]
                    x1 = word.geometry[1][0]
                    all_tokens.append({
                        "text": word.value, "y": y0 * h, "x": x0 * w,
                        "x2": x1 * w, "page": pi,
                    })

    all_tokens.sort(key=lambda t: (t["page"], t["y"], t["x"]))

    grouped_lines = []
    current_line = []
    current_y = -999
    current_page = -1
    for token in all_tokens:
        if token["page"] != current_page or abs(token["y"] - current_y) > 5:
            if current_line:
                grouped_lines.append(current_line)
            current_line = [token]
            current_y = token["y"]
            current_page = token["page"]
        else:
            current_line.append(token)
    if current_line:
        grouped_lines.append(current_line)

    lines = []
    for line_tokens in grouped_lines:
        line_tokens.sort(key=lambda t: t["x"])
        fused = []
        for token in line_tokens:
            if fused:
                prev = fused[-1]
                prev_ends_digit = prev["text"][-1].isdigit()
                cur_starts_num = token["text"][0].isdigit() or token["text"][0] in ",."
                distance = token["x"] - prev["x2"]
                if prev_ends_digit and cur_starts_num and abs(distance) < 20:
                    prev["text"] = prev["text"] + token["text"]
                    prev["x2"] = token["x2"]
                    continue
            fused.append(dict(token))
        lines.append(" ".join(t["text"] for t in fused))

    texte = "\n".join(lines)
    if not texte.strip():
        return {"fields": None, "error": "Aucun texte OCR", "installateur": None, "is_avoir": False}

    if fournisseur == "Auto-detect":
        installateur = detect_installateur(texte)
    elif fournisseur == "DEFAULT":
        installateur = "DEFAULT"
    else:
        installateur = fournisseur.lower()

    is_avoir = detect_avoir(texte)
    fields = extraire_champs(texte, installateur)
    if is_avoir and fields:
        fields = inverser_montants_avoir(fields)

    return {
        "fields": fields,
        "error": None,
        "installateur": installateur,
        "is_avoir": is_avoir,
    }


def process_batch_zip(zip_bytes: bytes) -> list[dict]:
    """Process a ZIP of invoices. Returns list of result dicts."""
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    valid_ext = {".pdf", ".png", ".jpg", ".jpeg"}
    file_names = [
        n for n in zf.namelist()
        if not n.startswith("__") and not n.endswith("/")
        and os.path.splitext(n)[1].lower() in valid_ext
    ]
    file_names.sort()

    results = []
    for fname in file_names:
        short_name = os.path.basename(fname)

        # Try pre-computed OCR first
        texte = get_ocr_text(short_name)
        if texte:
            installateur = detect_installateur(texte)
            is_avoir = detect_avoir(texte)
            fields = extraire_champs(texte, installateur)
            if is_avoir and fields:
                fields = inverser_montants_avoir(fields)
            if fields:
                results.append({
                    "filename": short_name,
                    "fields": fields,
                    "installateur": installateur or "DEFAULT",
                    "is_avoir": is_avoir,
                    "source": "OCR pre-calcule",
                })
                continue

        # Fallback: DocTR live
        try:
            file_bytes = zf.read(fname)
            suffix = os.path.splitext(fname)[1].lower()
            res = process_file_live(file_bytes, suffix, "Auto-detect")
            if res["fields"] and not res["error"]:
                results.append({
                    "filename": short_name,
                    "fields": res["fields"],
                    "installateur": res["installateur"] or "DEFAULT",
                    "is_avoir": res["is_avoir"],
                    "source": "DocTR live",
                })
            else:
                results.append({
                    "filename": short_name,
                    "fields": {k: None for k in ALL_FIELD_KEYS},
                    "installateur": "ERREUR",
                    "is_avoir": False,
                    "source": res["error"] or "Echec extraction",
                })
        except Exception as e:
            results.append({
                "filename": short_name,
                "fields": {k: None for k in ALL_FIELD_KEYS},
                "installateur": "ERREUR",
                "is_avoir": False,
                "source": str(e),
            })

    return results


# ─────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────

def export_excel_batch(results_list: list[dict]) -> bytes:
    """Generate styled Excel with one row per invoice. Returns bytes."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    all_labels = ["Nom du PDF", "Type"] + [lbl for _, lbl in FIELDS_LEFT + FIELDS_RIGHT] + ["Fournisseur"]
    rows = []
    for res in results_list:
        row = {"Nom du PDF": res["filename"]}
        row["Type"] = "AVOIR" if res.get("is_avoir") else "FACTURE"
        for key, label in FIELDS_LEFT + FIELDS_RIGHT:
            val = res["fields"].get(key)
            if val is None or val == "null":
                val = ""
            elif isinstance(val, list):
                val = " | ".join(str(v) for v in val)
            row[label] = str(val)
        row["Fournisseur"] = res.get("installateur", "").upper()
        rows.append(row)
    df = pd.DataFrame(rows, columns=all_labels)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extractions", startrow=2)
        ws = writer.sheets["Extractions"]

        nb_cols = len(all_labels)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
        title_cell = ws.cell(row=1, column=1, value=f"Extraction batch - {len(results_list)} factures")
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="1E3A5F")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        for col_idx in range(1, nb_cols + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 32

        val_font = Font(name="Calibri", size=11)
        val_fill_ok = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        val_fill_empty = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
        val_fill_avoir = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        val_fill_warn = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        val_font_warn = Font(name="Calibri", size=11, italic=True, color="E65100")
        label_to_key = {lbl: key for key, lbl in FIELDS_LEFT + FIELDS_RIGHT}

        for row_idx in range(4, 4 + len(results_list)):
            res = results_list[row_idx - 4]
            fields = res.get("fields") or {}
            a_verifier = fields.get("_a_verifier", [])
            is_avoir = res.get("is_avoir", False)

            for col_idx in range(1, nb_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = val_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                col_label = all_labels[col_idx - 1]
                field_key = label_to_key.get(col_label, "")

                if is_avoir:
                    cell.fill = val_fill_avoir
                    if col_idx == 1:
                        cell.font = Font(name="Calibri", size=11, bold=True)
                    elif not cell.value:
                        cell.value = "Non detecte"
                        cell.font = Font(name="Calibri", size=11, italic=True, color="C53030")
                elif col_idx == 1:
                    cell.font = Font(name="Calibri", size=11, bold=True)
                elif field_key in a_verifier:
                    cell.fill = val_fill_warn
                    cell.font = val_font_warn
                    if cell.value:
                        cell.value = str(cell.value) + " (a verifier)"
                elif not cell.value:
                    cell.fill = val_fill_empty
                    cell.value = "Non detecte"
                    cell.font = Font(name="Calibri", size=11, italic=True, color="C53030")
                else:
                    cell.fill = val_fill_ok
            ws.row_dimensions[row_idx].height = 26

        for col_idx, label in enumerate(all_labels, 1):
            max_len = len(label)
            for row_idx in range(4, 4 + len(results_list)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = max_len + 4

    return buf.getvalue()


def export_excel_multi_sheets(results_list: list[dict]) -> bytes:
    """Generate multi-sheet Excel (one per supplier + TOUTES_FACTURES). Returns bytes."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    columns = [
        "NOM_FICHIER", "TYPE", "NUMERO_FACTURE", "DATE_FACTURE",
        "MONTANT_HT", "TAUX_TVA", "MONTANT_TTC",
        "NOM_INSTALLATEUR", "COMMUNE_TRAVAUX", "CODE_POSTAL", "ADRESSE_TRAVAUX",
    ]
    field_keys = ALL_FIELD_KEYS

    def build_rows(items):
        rows = []
        for res in items:
            row = {"NOM_FICHIER": res["filename"]}
            row["TYPE"] = "AVOIR" if res.get("is_avoir") else "FACTURE"
            for col, key in zip(columns[2:], field_keys):
                val = res["fields"].get(key)
                if val is None or val == "null":
                    val = ""
                elif isinstance(val, list):
                    val = " | ".join(str(v) for v in val)
                row[col] = str(val)
            rows.append(row)
        return rows

    by_inst = {}
    for res in results_list:
        inst = res.get("installateur", "INCONNU").upper()
        by_inst.setdefault(inst, []).append(res)

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    val_font = Font(name="Calibri", size=11)
    err_font = Font(name="Calibri", size=11, italic=True, color="C53030")
    err_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
    ok_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    warn_font = Font(name="Calibri", size=11, italic=True, color="E65100")
    avoir_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )
    col_to_key = dict(zip(columns[2:], field_keys))

    def style_sheet(ws, rows_data, items):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=ws.title)
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="1E3A5F")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 32

        for row_idx in range(4, 4 + len(rows_data)):
            res = items[row_idx - 4]
            fields = res.get("fields") or {}
            a_verifier = fields.get("_a_verifier", [])
            is_avoir = res.get("is_avoir", False)

            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                col_name = columns[col_idx - 1]
                field_key = col_to_key.get(col_name, "")

                if is_avoir:
                    cell.fill = avoir_fill
                    if col_idx == 1:
                        cell.font = Font(name="Calibri", size=11, bold=True)
                    elif not cell.value:
                        cell.value = "Non detecte"
                        cell.font = err_font
                    else:
                        cell.font = val_font
                elif col_idx == 1:
                    cell.font = Font(name="Calibri", size=11, bold=True)
                elif field_key in a_verifier:
                    cell.fill = warn_fill
                    cell.font = warn_font
                    if cell.value:
                        cell.value = str(cell.value) + " (a verifier)"
                elif not cell.value:
                    cell.value = "Non detecte"
                    cell.font = err_font
                    cell.fill = err_fill
                else:
                    cell.font = val_font
                    cell.fill = ok_fill
            ws.row_dimensions[row_idx].height = 26

        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(col_name)
            for row_idx in range(4, 4 + len(rows_data)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = max_len + 4

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        all_rows = build_rows(results_list)
        df_all = pd.DataFrame(all_rows, columns=columns)
        df_all.to_excel(writer, index=False, sheet_name="TOUTES_FACTURES", startrow=2)
        style_sheet(writer.sheets["TOUTES_FACTURES"], all_rows, results_list)

        for inst_name in sorted(by_inst.keys()):
            sheet_name = inst_name[:31]
            inst_items = by_inst[inst_name]
            inst_rows = build_rows(inst_items)
            df_inst = pd.DataFrame(inst_rows, columns=columns)
            df_inst.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
            style_sheet(writer.sheets[sheet_name], inst_rows, inst_items)

    return buf.getvalue()


def get_fournisseurs_list() -> list[str]:
    """Return sorted list of known supplier names."""
    noms = [k for k in PROMPTS_INSTALLATEURS.keys() if k != "DEFAULT"]
    noms.sort()
    return noms
