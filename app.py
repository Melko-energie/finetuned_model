import os
import io
import json
import tempfile
import importlib.util
from pathlib import Path
import numpy as np
import pandas as pd
import streamlit as st
from PIL import Image
import fitz
import ollama
from doctr.models import ocr_predictor
from doctr.io import DocumentFile

PROJECT_ROOT = Path(__file__).resolve().parent
LOGO_PATH = PROJECT_ROOT / "assets" / "logo.png"

# Import pipeline from scripts/12_gemma2_smart.py
_spec_smart = importlib.util.spec_from_file_location(
    "gemma2_smart", str(PROJECT_ROOT / "scripts" / "12_gemma2_smart.py")
)
_mod_smart = importlib.util.module_from_spec(_spec_smart)
_spec_smart.loader.exec_module(_mod_smart)
smart_get_ocr_text = _mod_smart.get_ocr_text
smart_detect_installateur = _mod_smart.detect_installateur
smart_extraire_champs = _mod_smart.extraire_champs
smart_detect_avoir = _mod_smart.detect_avoir
smart_inverser_montants_avoir = _mod_smart.inverser_montants_avoir
SMART_INSTALLATEURS = _mod_smart.PROMPTS_INSTALLATEURS


@st.cache_resource
def load_ocr_model():
    return ocr_predictor(det_arch="db_resnet50", reco_arch="crnn_vgg16_bn", pretrained=True)


FIELDS_LEFT = [
    ("NUMERO_FACTURE",  "Numéro de facture"),
    ("DATE_FACTURE",    "Date de facture"),
    ("MONTANT_HT",      "Montant HT"),
    ("TAUX_TVA",        "Taux de TVA"),
    ("MONTANT_TTC",     "Montant TTC"),
]
FIELDS_RIGHT = [
    ("NOM_INSTALLATEUR", "Installateur"),
    ("COMMUNE_TRAVAUX",  "Commune"),
    ("CODE_POSTAL",      "Code Postal"),
    ("ADRESSE_TRAVAUX",  "Adresse des travaux"),
]

# ─────────────────────────────────────────────────────────────
# PAGE CONFIG & CSS
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Extraction Factures BTP",
    page_icon="🏗️",
    layout="wide",
)

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600&family=Syne:wght@700;800&display=swap');

  :root {
    --btp-blue:       #1E3A5F;
    --btp-blue-light: #2C5282;
    --bg-light:       #F7F8FA;
    --border:         #D1D5DB;
    --green-bg:       #E6F4EA;
    --green-border:   #34A853;
    --red-bg:         #FDE8E8;
    --red-border:     #C53030;
  }

  /* ── header ── */
  .main-header {
    background: linear-gradient(135deg, var(--btp-blue) 0%, var(--btp-blue-light) 100%);
    padding: 1.2rem 2rem;
    border-radius: 10px;
    margin-bottom: 1.5rem;
    color: white;
  }
  .main-header h1 { margin:0; font-size:1.8rem; font-weight:700; color:white; font-family:'Syne',sans-serif; }
  .main-header p  { margin:.25rem 0 0; font-size:.95rem; opacity:.85; color:#CBD5E0; }

  /* ── tabs ── */
  .stTabs [data-baseweb="tab-list"] {
    gap: 8px;
    border-bottom: 2px solid var(--border);
    margin-bottom: 1.5rem;
  }
  .stTabs [data-baseweb="tab"] {
    background: var(--bg-light);
    border-radius: 8px 8px 0 0;
    padding: 8px 20px;
    font-weight: 600;
    font-size: .9rem;
    border: 1px solid var(--border);
    border-bottom: none;
    color: #6B7280;
  }
  .stTabs [aria-selected="true"] {
    background: white !important;
    color: var(--btp-blue) !important;
    border-color: var(--btp-blue) !important;
  }

  /* ── upload zone ── */
  .upload-zone {
    border: 2px dashed var(--border);
    border-radius: 10px;
    padding: 1.5rem;
    text-align: center;
    background: var(--bg-light);
    margin-bottom: 1.5rem;
  }
  .upload-zone p { color:#6B7280; margin:.5rem 0 0; font-size:.9rem; }

  /* ── field cards ── */
  .field-card {
    border-radius: 8px;
    padding: .85rem 1rem;
    margin-bottom: .6rem;
    border-left: 4px solid;
  }
  .field-card.detected { background:var(--green-bg); border-left-color:var(--green-border); }
  .field-card.missing  { background:var(--red-bg);   border-left-color:var(--red-border);   }
  .field-label {
    font-size:.78rem; text-transform:uppercase;
    letter-spacing:.04em; color:#6B7280; margin-bottom:.15rem;
  }
  .field-value         { font-size:1rem; font-weight:600; color:#111827; word-break:break-word; }
  .field-value.missing { color:var(--red-border); font-weight:500; }

  /* ── section title ── */
  .section-title {
    font-size:1.05rem; font-weight:600; color:var(--btp-blue);
    margin-bottom:.75rem; padding-bottom:.4rem;
    border-bottom:2px solid var(--btp-blue); display:inline-block;
  }

  /* ── model badge ── */
  .model-badge {
    display:inline-flex; align-items:center; gap:6px;
    padding:4px 12px; border-radius:20px;
    font-size:.8rem; font-weight:600;
    margin-bottom:1rem;
  }
  .badge-gemma2   { background:rgba(124,58,237,.1);color:#7c3aed;           border:1px solid #7c3aed;           }
  .badge-smart    { background:rgba(234,88,12,.1); color:#ea580c;           border:1px solid #ea580c;           }

  /* ── stats row ── */
  .stats-row { display:flex; gap:1rem; margin-top:1rem; }
  .stat-box  {
    flex:1; background:var(--bg-light); border:1px solid var(--border);
    border-radius:8px; padding:.75rem 1rem; text-align:center;
  }
  .stat-box .num { font-size:1.4rem; font-weight:700; color:var(--btp-blue); }
  .stat-box .lbl { font-size:.78rem; color:#6B7280; }

  /* ── info box ── */
  .info-box {
    background: #EFF6FF; border:1px solid #BFDBFE;
    border-left:4px solid #3B82F6;
    border-radius:8px; padding:.85rem 1rem; margin-bottom:1rem;
    font-size:.88rem; color:#1E40AF; line-height:1.6;
  }
  .warn-box {
    background:#FFFBEB; border:1px solid #FDE68A;
    border-left:4px solid #F59E0B;
    border-radius:8px; padding:.85rem 1rem; margin-bottom:1rem;
    font-size:.88rem; color:#92400E; line-height:1.6;
  }

  /* ── footer ── */
  .app-footer {
    text-align:center; padding:1.5rem 0 .5rem;
    color:#9CA3AF; font-size:.82rem;
    border-top:1px solid var(--border); margin-top:2rem;
  }

  /* ── sidebar ── */
  section[data-testid="stSidebar"] { background:var(--bg-light); }
  .sidebar-section {
    background:white; border:1px solid var(--border);
    border-radius:8px; padding:.85rem 1rem; margin-bottom:.75rem;
  }
  .sidebar-section h4 { margin:0 0 .5rem; font-size:.9rem; color:var(--btp-blue); }
  .sidebar-row {
    display:flex; justify-content:space-between;
    font-size:.82rem; padding:.2rem 0; color:#374151;
  }
  .sidebar-row span:last-child { font-weight:600; }

  #MainMenu, footer, header { visibility:hidden; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────
# HELPERS COMMUNS
# ─────────────────────────────────────────────────────────────
def render_field_card(label: str, value):
    if value is None or value == "null":
        return (
            '<div class="field-card missing">'
            f'<div class="field-label">{label}</div>'
            '<div class="field-value missing">Non détecté</div>'
            '</div>'
        )
    if isinstance(value, list):
        value = " | ".join(str(v) for v in value)
    return (
        '<div class="field-card detected">'
        f'<div class="field-label">{label}</div>'
        f'<div class="field-value">{value}</div>'
        '</div>'
    )


def show_stats(fields, extra_stats=None):
    detected = sum(1 for k, _ in FIELDS_LEFT + FIELDS_RIGHT
                   if fields.get(k) and fields.get(k) != "null")
    total = len(FIELDS_LEFT) + len(FIELDS_RIGHT)
    html = (
        '<div class="stats-row">'
        f'<div class="stat-box"><div class="num">{detected}/{total}</div>'
        '<div class="lbl">Champs détectés</div></div>'
    )
    if extra_stats:
        for val, lbl in extra_stats:
            html += (
                f'<div class="stat-box"><div class="num">{val}</div>'
                f'<div class="lbl">{lbl}</div></div>'
            )
    html += '</div>'
    st.markdown(html, unsafe_allow_html=True)


def show_results(fields):
    c_left, c_right = st.columns(2)
    with c_left:
        for key, label in FIELDS_LEFT:
            st.markdown(render_field_card(label, fields.get(key)), unsafe_allow_html=True)
    with c_right:
        for key, label in FIELDS_RIGHT:
            st.markdown(render_field_card(label, fields.get(key)), unsafe_allow_html=True)


def show_preview(uploaded):
    suffix = os.path.splitext(uploaded.name)[1].lower()
    st.markdown('<div class="section-title">Aperçu du document</div>', unsafe_allow_html=True)
    if suffix == ".pdf":
        pdf_bytes = uploaded.read()
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        pix = doc[0].get_pixmap(matrix=fitz.Matrix(1.5, 1.5))
        preview = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        st.image(preview, use_container_width=True)
        doc.close()
        uploaded.seek(0)
    else:
        st.image(uploaded, use_container_width=True)
        uploaded.seek(0)


def export_excel(fields, filename, installateur="", is_avoir=False):
    """Wrapper : génère le même format Excel que le batch, avec une seule ligne."""
    return export_excel_batch([{
        "filename": filename,
        "fields": fields,
        "installateur": installateur or "",
        "is_avoir": is_avoir,
    }])


def export_excel_batch(results_list):
    """Génère un fichier Excel stylisé avec une ligne par facture."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    all_labels = ["Nom du PDF", "Type"] + [lbl for _, lbl in FIELDS_LEFT + FIELDS_RIGHT] + ["Fournisseur"]
    rows = []
    for res in results_list:
        row = {"Nom du PDF": res["filename"]}
        row["Type"] = "AVOIR" if res.get("is_avoir") else "FACTURE"
        for key, label in FIELDS_LEFT + FIELDS_RIGHT:
            val = res["fields"].get(key)
            if val is None or val == "null":
                val = ""
            elif isinstance(val, list):
                val = " | ".join(str(v) for v in val)
            row[label] = str(val)
        row["Fournisseur"] = res.get("installateur", "").upper()
        rows.append(row)
    df = pd.DataFrame(rows, columns=all_labels)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Extractions", startrow=2)
        ws = writer.sheets["Extractions"]

        # Titre
        nb_cols = len(all_labels)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nb_cols)
        title_cell = ws.cell(row=1, column=1, value=f"Extraction batch — {len(results_list)} factures")
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="1E3A5F")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        # Style en-têtes (ligne 3)
        for col_idx in range(1, nb_cols + 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 32

        # Style lignes de données
        val_font = Font(name="Calibri", size=11)
        val_fill_ok = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
        val_fill_empty = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
        val_fill_alt = PatternFill(start_color="F7F8FA", end_color="F7F8FA", fill_type="solid")
        val_fill_warn = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        val_font_warn = Font(name="Calibri", size=11, italic=True, color="E65100")
        val_fill_avoir = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")

        # Mapper label → clé pour identifier les champs à vérifier
        label_to_key = {lbl: key for key, lbl in FIELDS_LEFT + FIELDS_RIGHT}

        for row_idx in range(4, 4 + len(results_list)):
            res = results_list[row_idx - 4]
            fields = res.get("fields") or {}
            a_verifier = fields.get("_a_verifier", [])
            is_avoir = res.get("is_avoir", False)

            for col_idx in range(1, nb_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = val_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                col_label = all_labels[col_idx - 1]
                field_key = label_to_key.get(col_label, "")

                if is_avoir:
                    # Ligne avoir → fond jaune pour toutes les cellules
                    cell.fill = val_fill_avoir
                    if col_idx == 1:
                        cell.font = Font(name="Calibri", size=11, bold=True)
                    elif not cell.value:
                        cell.value = "Non detecte"
                        cell.font = Font(name="Calibri", size=11, italic=True, color="C53030")
                elif col_idx == 1:
                    # Colonne nom du PDF : fond alterné
                    cell.font = Font(name="Calibri", size=11, bold=True)
                    if (row_idx - 4) % 2 == 1:
                        cell.fill = val_fill_alt
                elif field_key in a_verifier:
                    # Champ incohérent → orange "à vérifier"
                    cell.fill = val_fill_warn
                    cell.font = val_font_warn
                    if cell.value:
                        cell.value = str(cell.value) + " (a verifier)"
                elif not cell.value:
                    cell.fill = val_fill_empty
                    cell.value = "Non detecte"
                    cell.font = Font(name="Calibri", size=11, italic=True, color="C53030")
                else:
                    cell.fill = val_fill_ok
            ws.row_dimensions[row_idx].height = 26

        # Largeur colonnes auto
        for col_idx, label in enumerate(all_labels, 1):
            max_len = len(label)
            for row_idx in range(4, 4 + len(results_list)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = max_len + 4

    return buf.getvalue()


def process_single_file(file_bytes, suffix, choix_fournisseur):
    """OCR DocTR en mémoire + extraction Gemma2. Retourne (fields, installateur) ou (None, error)."""
    # Convertir en images PIL
    pil_images = []
    if suffix == ".pdf":
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        for page in doc:
            pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            pil_images.append(img)
        doc.close()
    else:
        pil_images.append(Image.open(io.BytesIO(file_bytes)).convert("RGB"))

    nb_pages = len(pil_images)

    # OCR DocTR
    model_ocr = load_ocr_model()
    np_images = [np.array(img) for img in pil_images]
    result_ocr = model_ocr(np_images)

    # Chercher la page des montants
    MONTANT_KEYWORDS = ["ht", "tva", "ttc", "total", "net", "payer", "montant"]
    if nb_pages == 1:
        pages_indices = [0]
    else:
        best_page = nb_pages - 1
        best_score = 0
        for pi in range(1, nb_pages):
            words = [
                word.value.lower()
                for block in result_ocr.pages[pi].blocks
                for line in block.lines
                for word in line.words
            ]
            score = sum(1 for w in words for kw in MONTANT_KEYWORDS if kw in w)
            if score > best_score or (score == best_score and pi > best_page):
                best_score = score
                best_page = pi
        pages_indices = [0] if best_page == 0 else [0, best_page]

    # Reconstruire le texte
    all_tokens = []
    for pi in pages_indices:
        page = result_ocr.pages[pi]
        h, w = pil_images[pi].height, pil_images[pi].width
        for block in page.blocks:
            for line in block.lines:
                for word in line.words:
                    x0, y0 = word.geometry[0]
                    x1 = word.geometry[1][0]
                    all_tokens.append({"text": word.value, "y": y0 * h, "x": x0 * w, "x2": x1 * w, "page": pi})

    # Regrouper par lignes approximatives (tolérance Y de 5 pixels)
    all_tokens.sort(key=lambda t: (t["page"], t["y"], t["x"]))

    grouped_lines = []
    current_line = []
    current_y = -999
    current_page = -1
    for token in all_tokens:
        if token["page"] != current_page or abs(token["y"] - current_y) > 5:
            if current_line:
                grouped_lines.append(current_line)
            current_line = [token]
            current_y = token["y"]
            current_page = token["page"]
        else:
            current_line.append(token)
    if current_line:
        grouped_lines.append(current_line)

    # Trier chaque ligne par X, puis fusionner les tokens numériques adjacents
    lines = []
    for line_tokens in grouped_lines:
        line_tokens.sort(key=lambda t: t["x"])

        fused = []
        for token in line_tokens:
            if fused:
                prev = fused[-1]
                prev_ends_digit = prev["text"][-1].isdigit()
                cur_starts_num = token["text"][0].isdigit() or token["text"][0] in ",."
                distance = token["x"] - prev["x2"]
                if prev_ends_digit and cur_starts_num and abs(distance) < 20:
                    prev["text"] = prev["text"] + token["text"]
                    prev["x2"] = token["x2"]
                    continue
            fused.append(dict(token))

        lines.append(" ".join(t["text"] for t in fused))

    texte = "\n".join(lines)
    if not texte.strip():
        return None, "Aucun texte OCR", None

    # Détection fournisseur + extraction
    if choix_fournisseur == "Auto-detect":
        installateur = smart_detect_installateur(texte)
    elif choix_fournisseur == "DEFAULT (generique)":
        installateur = "DEFAULT"
    else:
        installateur = choix_fournisseur.lower()

    is_avoir = smart_detect_avoir(texte)
    fields = smart_extraire_champs(texte, installateur)

    if is_avoir and fields:
        fields = smart_inverser_montants_avoir(fields)

    return fields, None, installateur, is_avoir


def clean_json(raw):
    raw = raw.strip()
    if "```json" in raw:
        raw = raw.split("```json")[1].split("```")[0].strip()
    elif "```" in raw:
        raw = raw.split("```")[1].split("```")[0].strip()
    return raw


# ─────────────────────────────────────────────────────────────
# HELPERS GEMMA2 TEXTE
# ─────────────────────────────────────────────────────────────
PROMPT_TEXTE = """Tu es un expert comptable spécialisé dans les factures BTP françaises.

RÈGLES :
- NUMERO_FACTURE : numéro COMPLET après "N°", "Facture N°", "Réf". Garde tous les caractères.
- DATE_FACTURE : format JJ/MM/AAAA.
- MONTANT_HT : montant AVANT TVA. Format "358.83". Cherche "Total HT", "Net HT".
- TAUX_TVA : uniquement le pourcentage. Exemple "20%".
- MONTANT_TTC : montant FINAL avec TVA. Cherche "Total TTC", "Net à payer".
- NOM_INSTALLATEUR : nom COMPLET de l'entreprise émettrice (pas le client).
- COMMUNE_TRAVAUX : ville du chantier.
- CODE_POSTAL : code postal à 5 chiffres.
- ADRESSE_TRAVAUX : adresse complète du chantier.

TEXTE DE LA FACTURE :
{texte}

Réponds UNIQUEMENT en JSON valide. Aucun texte avant ou après.
{{"NUMERO_FACTURE":null,"DATE_FACTURE":null,"MONTANT_HT":null,"TAUX_TVA":null,"MONTANT_TTC":null,"NOM_INSTALLATEUR":null,"COMMUNE_TRAVAUX":null,"CODE_POSTAL":null,"ADRESSE_TRAVAUX":null}}"""


def extract_gemma2(original_filename):
    """Extraction texte avec gemma2:9b — utilise l'OCR DocTR existant"""
    ocr_dir = PROJECT_ROOT / "data" / "ocr_texts"
    base_name = os.path.splitext(original_filename)[0]
    base_name = base_name.replace("_page0", "").replace("_page_001", "")

    texte = ""
    for root, dirs, files in os.walk(str(ocr_dir)):
        for f in files:
            if base_name.lower() in f.lower():
                try:
                    with open(os.path.join(root, f), "r", encoding="utf-8") as fp:
                        data = json.load(fp)
                        pages_ocr = data.get("pages", [])
                        for page in pages_ocr:
                            if isinstance(page, list):
                                lignes = {}
                                for token in page:
                                    if isinstance(token, dict):
                                        text = token.get("text", "").strip()
                                        if not text:
                                            continue
                                        bbox = token.get("bbox", [0,0,0,0])
                                        y = round(bbox[1] / 10) * 10
                                        if y not in lignes:
                                            lignes[y] = []
                                        lignes[y].append((bbox[0], text))
                                for y in sorted(lignes.keys()):
                                    tokens_ligne = sorted(lignes[y], key=lambda t: t[0])
                                    texte += "  ".join(t[1] for t in tokens_ligne) + "\n"
                except Exception:
                    pass
                break

    if not texte:
        return None, "OCR introuvable pour cette facture", False

    is_avoir = smart_detect_avoir(texte)

    prompt = PROMPT_TEXTE.format(texte=texte[:3000])
    response = ollama.chat(
        model="gemma2:9b",
        options={"temperature": 0, "seed": 42},
        messages=[{"role":"user","content":prompt}]
    )

    try:
        result = json.loads(clean_json(response["message"]["content"]))
        if is_avoir and result:
            result = smart_inverser_montants_avoir(result)
        return result, None, is_avoir
    except json.JSONDecodeError:
        return None, "Réponse JSON invalide", False


# ─────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────
with st.sidebar:
    if LOGO_PATH.exists():
        st.image(str(LOGO_PATH), width=150)
    st.markdown("")

    st.markdown(
        '<div class="sidebar-section">'
        "<h4>📋 Projet</h4>"
        '<div class="sidebar-row"><span>OCR</span><span>DocTR</span></div>'
        '<div class="sidebar-row"><span>Modèle</span><span>Gemma2:9b</span></div>'
        '<div class="sidebar-row"><span>Version</span><span>3.0.0</span></div>'
        "</div>",
        unsafe_allow_html=True,
    )

    st.markdown(
        '<div class="sidebar-section">'
        "<h4>🤖 Modèles disponibles</h4>"
        '<div class="sidebar-row"><span>Gemma2:9b</span><span style="color:#7c3aed">Texte</span></div>'
        '<div class="sidebar-row"><span>Gemma2 Smart</span><span style="color:#ea580c">Par fournisseur</span></div>'
        "</div>",
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────
if LOGO_PATH.exists():
    col_logo, col_title = st.columns([1, 4])
    with col_logo:
        st.image(str(LOGO_PATH), width=120)
    with col_title:
        st.markdown(
            '<div class="main-header">'
            '<h1>🏗️ Extraction de Factures BTP</h1>'
            '<p>Gemma2:9b Texte · Gemma2 Smart par fournisseur</p>'
            '</div>',
            unsafe_allow_html=True,
        )
else:
    st.markdown(
        '<div class="main-header">'
        '<h1>🏗️ Extraction de Factures BTP</h1>'
        '<p>Gemma2:9b Texte · Gemma2 Smart par fournisseur</p>'
        '</div>',
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────────────────────
# ONGLETS
# ─────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs([
    "📝 Gemma2:9b — Texte",
    "🎯 Gemma2 Smart — Par fournisseur",
    "🔍 Nouvelle Facture — OCR + Extraction",
    "📦 Traitement par lot — ZIP",
])


# ═══════════════════════════════════════════════════
# ONGLET 1 — GEMMA2:9b TEXTE
# ═══════════════════════════════════════════════════
with tab1:
    st.markdown(
        '<span class="model-badge badge-gemma2">📝 Gemma2:9b — Extraction textuelle</span>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="warn-box"><strong>Gemma2:9b</strong> utilise le texte OCR extrait par DocTR '
        '(script 01). Il ne voit pas l\'image directement. Reconstruction des lignes par '
        'position Y pour préserver la structure des tableaux. '
        '<strong>Attention</strong> : nécessite que l\'OCR ait déjà été extrait via script 01.</div>',
        unsafe_allow_html=True,
    )

    st.markdown(
        '<div class="upload-zone">'
        '<p><strong>Déposez votre facture ici</strong></p>'
        '<p>Formats acceptés : PNG, JPG, PDF — OCR DocTR requis</p>'
        '</div>',
        unsafe_allow_html=True,
    )

    uploaded1 = st.file_uploader(
        "Importer (Gemma2)",
        type=["png","jpg","jpeg","pdf"],
        label_visibility="collapsed",
        key="upload_gemma2",
    )

    if uploaded1:
        col_prev1, col_res1 = st.columns([1,1], gap="large")

        with col_prev1:
            show_preview(uploaded1)

        with st.spinner("Gemma2:9b analyse le texte OCR..."):
            fields1, error1, is_avoir1 = extract_gemma2(original_filename=uploaded1.name)

        with col_res1:
            st.markdown('<div class="section-title">Champs extraits</div>', unsafe_allow_html=True)

            if error1:
                st.error(f"  {error1}")
                st.info(
                    "Vérifiez que l'OCR a bien été extrait via `python scripts/01_ocr_extraction.py` "
                    "et que le PDF source est dans `data/raw_pdfs/`"
                )
            elif fields1:
                show_results(fields1)
                show_stats(fields1)

        if fields1 and not error1:
            with st.expander("JSON brut"):
                st.json(fields1)
            base_name = os.path.splitext(uploaded1.name)[0]
            st.download_button(
                "Exporter en Excel",
                data=export_excel(fields1, uploaded1.name, "gemma2-texte", is_avoir=is_avoir1),
                file_name=f"{base_name}_extraction.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel_tab1",
            )


# ═══════════════════════════════════════════════════
# ONGLET 2 — GEMMA2 SMART (prompts par fournisseur)
# ═══════════════════════════════════════════════════
with tab2:
    st.markdown(
        '<span class="model-badge badge-smart">🎯 Gemma2:9b — Prompts spécialisés par fournisseur</span>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="info-box"><strong>Gemma2 Smart</strong> détecte automatiquement le fournisseur '
        'à partir du texte OCR et utilise un <strong>prompt spécialisé</strong> adapté à ses factures '
        '(format n° facture, taux TVA habituel, emplacement des montants). '
        '17 fournisseurs connus + prompt générique pour les inconnus. '
        '<strong>Nécessite</strong> que l\'OCR ait été extrait via script 01.</div>',
        unsafe_allow_html=True,
    )

    # Liste déroulante des fournisseurs
    noms_installateurs = [k for k in SMART_INSTALLATEURS.keys() if k != "DEFAULT"]
    noms_installateurs.sort()
    options_dropdown = ["Auto-detect"] + [n.upper() for n in noms_installateurs] + ["DEFAULT (generique)"]

    choix_installateur = st.selectbox(
        "Choisir le fournisseur",
        options_dropdown,
        index=0,
        help="Selectionnez le fournisseur manuellement si la facture contient plusieurs noms d'installateurs",
        key="select_installateur_smart",
    )

    st.markdown(
        '<div class="upload-zone">'
        '<p><strong>Deposez votre facture ici</strong></p>'
        '<p>Formats acceptes : PNG, JPG, PDF — OCR DocTR requis</p>'
        '</div>',
        unsafe_allow_html=True,
    )

    uploaded2 = st.file_uploader(
        "Importer (Gemma2 Smart)",
        type=["png","jpg","jpeg","pdf"],
        label_visibility="collapsed",
        key="upload_gemma2_smart",
    )

    if uploaded2:
        col_prev2, col_res2 = st.columns([1,1], gap="large")

        with col_prev2:
            show_preview(uploaded2)

        with st.spinner("Gemma2 Smart — detection fournisseur et extraction..."):
            texte_ocr = smart_get_ocr_text(uploaded2.name)

            if texte_ocr is None:
                installateur_detecte = None
                fields2 = None
                error2 = "OCR introuvable pour cette facture"
                is_avoir2 = False
            else:
                # Choix manuel ou auto-detect
                if choix_installateur == "Auto-detect":
                    installateur_detecte = smart_detect_installateur(texte_ocr)
                    mode_detection = "auto"
                elif choix_installateur == "DEFAULT (generique)":
                    installateur_detecte = "DEFAULT"
                    mode_detection = "manuel"
                else:
                    installateur_detecte = choix_installateur.lower()
                    mode_detection = "manuel"

                is_avoir2 = smart_detect_avoir(texte_ocr)
                fields2 = smart_extraire_champs(texte_ocr, installateur_detecte)
                if is_avoir2 and fields2:
                    fields2 = smart_inverser_montants_avoir(fields2)
                error2 = None

        with col_res2:
            st.markdown('<div class="section-title">Champs extraits</div>', unsafe_allow_html=True)

            if error2:
                st.error(f"  {error2}")
                st.info(
                    "Verifiez que l'OCR a bien ete extrait via `python scripts/01_ocr_extraction.py` "
                    "et que le PDF source est dans `data/raw_pdfs/`"
                )
            elif fields2:
                # Afficher le fournisseur detecte/choisi
                if installateur_detecte and installateur_detecte != "DEFAULT":
                    label_mode = "detecte automatiquement" if mode_detection == "auto" else "choisi manuellement"
                    st.markdown(
                        f'<div style="background:#FFF7ED;border:1px solid #FDBA74;border-left:4px solid #ea580c;'
                        f'border-radius:8px;padding:.75rem 1rem;margin-bottom:1rem;font-size:.88rem;color:#9A3412;">'
                        f'🎯 Fournisseur {label_mode} : <strong>{installateur_detecte.upper()}</strong> '
                        f'— prompt specialise utilise</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        '<div style="background:#F3F4F6;border:1px solid #D1D5DB;border-left:4px solid #6B7280;'
                        'border-radius:8px;padding:.75rem 1rem;margin-bottom:1rem;font-size:.88rem;color:#374151;">'
                        'Fournisseur non reconnu — prompt generique utilise</div>',
                        unsafe_allow_html=True,
                    )

                show_results(fields2)
                show_stats(fields2, [
                    (installateur_detecte.upper() if installateur_detecte else "?", "Fournisseur"),
                ])

        if fields2 and not error2:
            with st.expander("JSON brut"):
                st.json(fields2)
            base_name2 = os.path.splitext(uploaded2.name)[0]
            st.download_button(
                "Exporter en Excel",
                data=export_excel(fields2, uploaded2.name, installateur_detecte or "", is_avoir=is_avoir2),
                file_name=f"{base_name2}_extraction_smart.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel_tab2",
            )


# ═══════════════════════════════════════════════════
# ONGLET 3 — NOUVELLE FACTURE — OCR + EXTRACTION
# ═══════════════════════════════════════════════════
with tab3:
    st.markdown(
        '<span class="model-badge badge-smart">🔍 OCR DocTR + Gemma2:9b — Pipeline complet en mémoire</span>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="info-box"><strong>Pipeline complet</strong> : uploadez directement un PDF, une image, '
        'ou un <strong>dossier ZIP</strong> contenant plusieurs factures. '
        'DocTR effectue l\'OCR en mémoire, puis Gemma2 Smart détecte le fournisseur et extrait les champs '
        'avec le prompt spécialisé. <strong>Aucun fichier intermédiaire</strong> n\'est sauvegardé sur disque.</div>',
        unsafe_allow_html=True,
    )

    # Dropdown fournisseur
    noms_inst3 = [k for k in SMART_INSTALLATEURS.keys() if k != "DEFAULT"]
    noms_inst3.sort()
    options3 = ["Auto-detect"] + [n.upper() for n in noms_inst3] + ["DEFAULT (generique)"]

    choix3 = st.selectbox(
        "Forcer le fournisseur (optionnel)",
        options3,
        index=0,
        help="Laissez Auto-detect ou forcez manuellement si la détection est mauvaise",
        key="select_installateur_ocr",
    )

    mode_input = st.radio(
        "Mode d'import",
        ["Fichier unique (PDF/image)", "Dossier ZIP (plusieurs factures)"],
        horizontal=True,
        key="mode_input_tab3",
    )

    # ── MODE FICHIER UNIQUE ──
    if mode_input == "Fichier unique (PDF/image)":
        st.markdown(
            '<div class="upload-zone">'
            '<p><strong>Déposez votre facture ici</strong></p>'
            '<p>Formats acceptés : PNG, JPG, PDF</p>'
            '</div>',
            unsafe_allow_html=True,
        )

        uploaded3 = st.file_uploader(
            "Importer (OCR + Extraction)",
            type=["png", "jpg", "jpeg", "pdf"],
            label_visibility="collapsed",
            key="upload_ocr_extract",
        )

        if uploaded3:
            col_prev3, col_res3 = st.columns([1, 1], gap="large")

            with col_prev3:
                show_preview(uploaded3)

            progress = st.progress(0, text="Extraction OCR...")
            file_bytes = uploaded3.read()
            suffix = os.path.splitext(uploaded3.name)[1].lower()

            progress.progress(30, text="Extraction OCR... (DocTR)")
            fields3, error3, installateur3, is_avoir3 = process_single_file(file_bytes, suffix, choix3)
            progress.progress(100, text="Terminé !")
            progress.empty()

            with col_res3:
                st.markdown('<div class="section-title">Champs extraits</div>', unsafe_allow_html=True)

                if error3:
                    st.error(error3)
                elif fields3:
                    # Badge fournisseur
                    if installateur3 and installateur3 != "DEFAULT":
                        mode3 = "détecté automatiquement" if choix3 == "Auto-detect" else "choisi manuellement"
                        st.markdown(
                            f'<div style="background:#FFF7ED;border:1px solid #FDBA74;border-left:4px solid #ea580c;'
                            f'border-radius:8px;padding:.75rem 1rem;margin-bottom:1rem;font-size:.88rem;color:#9A3412;">'
                            f'🎯 Fournisseur {mode3} : <strong>{installateur3.upper()}</strong> '
                            f'— prompt spécialisé utilisé</div>',
                            unsafe_allow_html=True,
                        )
                    else:
                        st.markdown(
                            '<div style="background:#F3F4F6;border:1px solid #D1D5DB;border-left:4px solid #6B7280;'
                            'border-radius:8px;padding:.75rem 1rem;margin-bottom:1rem;font-size:.88rem;color:#374151;">'
                            'Fournisseur non reconnu — prompt générique utilisé</div>',
                            unsafe_allow_html=True,
                        )

                    show_results(fields3)
                    show_stats(fields3, [
                        (installateur3.upper() if installateur3 else "?", "Fournisseur"),
                    ])

            if fields3 and not error3:
                with st.expander("JSON brut"):
                    st.json(fields3)
                base_name3 = os.path.splitext(uploaded3.name)[0]
                st.download_button(
                    "Exporter en Excel",
                    data=export_excel(fields3, uploaded3.name, installateur3 or "", is_avoir=is_avoir3),
                    file_name=f"{base_name3}_extraction_ocr.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_excel_tab3",
                )

    # ── MODE ZIP (BATCH) ──
    else:
        import zipfile

        st.markdown(
            '<div class="upload-zone">'
            '<p><strong>Déposez votre dossier ZIP ici</strong></p>'
            '<p>Le ZIP doit contenir des fichiers PDF, PNG ou JPG</p>'
            '</div>',
            unsafe_allow_html=True,
        )

        uploaded_zip = st.file_uploader(
            "Importer ZIP",
            type=["zip"],
            label_visibility="collapsed",
            key="upload_zip_batch",
        )

        if uploaded_zip:
            zip_bytes = uploaded_zip.read()
            zf = zipfile.ZipFile(io.BytesIO(zip_bytes))

            # Filtrer les fichiers valides (ignorer dossiers, __MACOSX, etc.)
            valid_ext = {".pdf", ".png", ".jpg", ".jpeg"}
            file_names = [
                n for n in zf.namelist()
                if not n.startswith("__") and os.path.splitext(n)[1].lower() in valid_ext
            ]
            file_names.sort()

            if not file_names:
                st.error("Aucun fichier PDF/PNG/JPG trouvé dans le ZIP.")
            else:
                st.markdown(
                    f'<div class="info-box">📦 <strong>{len(file_names)} factures</strong> '
                    f'détectées dans le ZIP</div>',
                    unsafe_allow_html=True,
                )

                batch_results = []
                progress_bar = st.progress(0, text="Traitement batch...")

                for i, fname in enumerate(file_names):
                    short_name = os.path.basename(fname)
                    progress_bar.progress(
                        (i) / len(file_names),
                        text=f"Facture {i+1}/{len(file_names)} : {short_name}",
                    )

                    file_bytes = zf.read(fname)
                    suffix = os.path.splitext(fname)[1].lower()
                    fields, error, installateur, is_avoir = process_single_file(file_bytes, suffix, choix3)

                    if fields and not error:
                        batch_results.append({
                            "filename": short_name,
                            "fields": fields,
                            "installateur": installateur or "DEFAULT",
                            "is_avoir": is_avoir,
                        })
                    else:
                        batch_results.append({
                            "filename": short_name,
                            "fields": {k: None for k, _ in FIELDS_LEFT + FIELDS_RIGHT},
                            "installateur": "ERREUR",
                            "is_avoir": False,
                        })

                progress_bar.progress(1.0, text=f"Terminé — {len(batch_results)} factures traitées !")
                progress_bar.empty()

                # Afficher les résultats de chaque facture
                for res in batch_results:
                    is_error = res["installateur"] == "ERREUR"
                    with st.expander(
                        f"{'❌' if is_error else '✅'} {res['filename']} — {res['installateur'].upper()}",
                        expanded=False,
                    ):
                        if is_error:
                            st.error("Extraction échouée pour cette facture.")
                        else:
                            show_results(res["fields"])

                # Bouton export Excel batch
                st.markdown("---")
                nb_ok = sum(1 for r in batch_results if r["installateur"] != "ERREUR")
                st.markdown(
                    f'<div class="stats-row">'
                    f'<div class="stat-box"><div class="num">{nb_ok}/{len(batch_results)}</div>'
                    f'<div class="lbl">Factures extraites</div></div>'
                    f'<div class="stat-box"><div class="num">{len(batch_results) - nb_ok}</div>'
                    f'<div class="lbl">Erreurs</div></div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                zip_base = os.path.splitext(uploaded_zip.name)[0]
                st.download_button(
                    f"Exporter les {len(batch_results)} résultats en Excel",
                    data=export_excel_batch(batch_results),
                    file_name=f"{zip_base}_batch_extraction.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_excel_batch",
                )


# ═══════════════════════════════════════════════════
# ONGLET 4 — TRAITEMENT PAR LOT — ZIP
# ═══════════════════════════════════════════════════

def export_excel_multi_sheets(results_list):
    """Excel multi-feuilles : une feuille par installateur + TOUTES_FACTURES."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    columns = [
        "NOM_FICHIER", "TYPE", "NUMERO_FACTURE", "DATE_FACTURE",
        "MONTANT_HT", "TAUX_TVA", "MONTANT_TTC",
        "NOM_INSTALLATEUR", "COMMUNE_TRAVAUX", "CODE_POSTAL", "ADRESSE_TRAVAUX",
    ]
    field_keys = [
        "NUMERO_FACTURE", "DATE_FACTURE", "MONTANT_HT", "TAUX_TVA", "MONTANT_TTC",
        "NOM_INSTALLATEUR", "COMMUNE_TRAVAUX", "CODE_POSTAL", "ADRESSE_TRAVAUX",
    ]

    def build_rows(items):
        rows = []
        for res in items:
            row = {"NOM_FICHIER": res["filename"]}
            row["TYPE"] = "AVOIR" if res.get("is_avoir") else "FACTURE"
            for col, key in zip(columns[2:], field_keys):
                val = res["fields"].get(key)
                if val is None or val == "null":
                    val = ""
                elif isinstance(val, list):
                    val = " | ".join(str(v) for v in val)
                row[col] = str(val)
            rows.append(row)
        return rows

    # Grouper par installateur
    by_inst = {}
    for res in results_list:
        inst = res.get("installateur", "INCONNU").upper()
        by_inst.setdefault(inst, []).append(res)

    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    val_font = Font(name="Calibri", size=11)
    err_font = Font(name="Calibri", size=11, italic=True, color="C53030")
    err_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid")
    ok_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
    warn_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
    warn_font = Font(name="Calibri", size=11, italic=True, color="E65100")
    avoir_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Mapper colonne → clé pour les champs à vérifier
    col_to_key = dict(zip(columns[2:], field_keys))

    def style_sheet(ws, rows_data, items):
        # Titre
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
        title_cell = ws.cell(row=1, column=1, value=ws.title)
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="1E3A5F")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # En-tetes (ligne 3)
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border
        ws.row_dimensions[3].height = 32

        # Valeurs
        for row_idx in range(4, 4 + len(rows_data)):
            res = items[row_idx - 4]
            fields = res.get("fields") or {}
            a_verifier = fields.get("_a_verifier", [])
            is_avoir = res.get("is_avoir", False)

            for col_idx in range(1, len(columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                col_name = columns[col_idx - 1]
                field_key = col_to_key.get(col_name, "")

                if is_avoir:
                    # Ligne avoir → fond jaune pour toutes les cellules
                    cell.fill = avoir_fill
                    if col_idx == 1:
                        cell.font = Font(name="Calibri", size=11, bold=True)
                    elif not cell.value:
                        cell.value = "Non detecte"
                        cell.font = err_font
                    else:
                        cell.font = val_font
                elif col_idx == 1:
                    cell.font = Font(name="Calibri", size=11, bold=True)
                elif field_key in a_verifier:
                    cell.fill = warn_fill
                    cell.font = warn_font
                    if cell.value:
                        cell.value = str(cell.value) + " (a verifier)"
                elif not cell.value:
                    cell.value = "Non detecte"
                    cell.font = err_font
                    cell.fill = err_fill
                else:
                    cell.font = val_font
                    cell.fill = ok_fill
            ws.row_dimensions[row_idx].height = 26

        # Largeur auto
        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(col_name)
            for row_idx in range(4, 4 + len(rows_data)):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, len(val))
            ws.column_dimensions[ws.cell(row=3, column=col_idx).column_letter].width = max_len + 4

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Feuille TOUTES_FACTURES
        all_rows = build_rows(results_list)
        df_all = pd.DataFrame(all_rows, columns=columns)
        df_all.to_excel(writer, index=False, sheet_name="TOUTES_FACTURES", startrow=2)
        style_sheet(writer.sheets["TOUTES_FACTURES"], all_rows, results_list)

        # Une feuille par installateur
        for inst_name in sorted(by_inst.keys()):
            sheet_name = inst_name[:31]  # Excel limite à 31 chars
            inst_items = by_inst[inst_name]
            inst_rows = build_rows(inst_items)
            df_inst = pd.DataFrame(inst_rows, columns=columns)
            df_inst.to_excel(writer, index=False, sheet_name=sheet_name, startrow=2)
            style_sheet(writer.sheets[sheet_name], inst_rows, inst_items)

    return buf.getvalue()


with tab4:
    st.markdown(
        '<span class="model-badge badge-smart">📦 Traitement par lot — ZIP multi-factures</span>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="info-box"><strong>Mode batch</strong> : uploadez un ZIP contenant vos factures PDF. '
        'Le système cherche d\'abord l\'OCR pre-calcule (data/ocr_texts/), '
        'sinon fait l\'OCR DocTR en memoire. Chaque facture est traitee avec le prompt '
        'specialise du fournisseur detecte. Export Excel multi-feuilles (une par installateur).</div>',
        unsafe_allow_html=True,
    )

    import zipfile

    st.markdown(
        '<div class="upload-zone">'
        '<p><strong>Deposez votre dossier ZIP ici</strong></p>'
        '<p>Le ZIP peut contenir des PDFs a la racine et/ou dans des sous-dossiers</p>'
        '</div>',
        unsafe_allow_html=True,
    )

    uploaded_zip4 = st.file_uploader(
        "Importer ZIP (batch)",
        type=["zip"],
        label_visibility="collapsed",
        key="upload_zip_tab4",
    )

    if uploaded_zip4:
        zip_bytes4 = uploaded_zip4.read()
        zf4 = zipfile.ZipFile(io.BytesIO(zip_bytes4))

        valid_ext = {".pdf", ".png", ".jpg", ".jpeg"}
        file_names4 = [
            n for n in zf4.namelist()
            if not n.startswith("__") and not n.endswith("/")
            and os.path.splitext(n)[1].lower() in valid_ext
        ]
        file_names4.sort()

        if not file_names4:
            st.error("Aucun fichier PDF/PNG/JPG trouve dans le ZIP.")
        else:
            st.markdown(
                f'<div class="info-box">📦 <strong>{len(file_names4)} factures</strong> '
                f'detectees dans le ZIP</div>',
                unsafe_allow_html=True,
            )

            batch_results4 = []
            progress4 = st.progress(0, text="Traitement batch...")

            for i, fname in enumerate(file_names4):
                short_name = os.path.basename(fname)
                progress4.progress(
                    i / len(file_names4),
                    text=f"Traitement facture {i+1}/{len(file_names4)} : {short_name}",
                )

                # Essayer OCR pre-calcule d'abord
                texte = smart_get_ocr_text(short_name)

                if texte:
                    # OCR pre-calcule trouve
                    installateur = smart_detect_installateur(texte)
                    is_avoir = smart_detect_avoir(texte)
                    fields = smart_extraire_champs(texte, installateur)
                    if is_avoir and fields:
                        fields = smart_inverser_montants_avoir(fields)
                    if fields:
                        batch_results4.append({
                            "filename": short_name,
                            "fields": fields,
                            "installateur": installateur or "DEFAULT",
                            "is_avoir": is_avoir,
                            "source": "OCR pre-calcule",
                        })
                        continue

                # Fallback : DocTR en memoire
                try:
                    file_bytes4 = zf4.read(fname)
                    suffix4 = os.path.splitext(fname)[1].lower()
                    fields, error, installateur, is_avoir = process_single_file(
                        file_bytes4, suffix4, "Auto-detect"
                    )
                    if fields and not error:
                        batch_results4.append({
                            "filename": short_name,
                            "fields": fields,
                            "installateur": installateur or "DEFAULT",
                            "is_avoir": is_avoir,
                            "source": "DocTR live",
                        })
                    else:
                        batch_results4.append({
                            "filename": short_name,
                            "fields": {k: None for k, _ in FIELDS_LEFT + FIELDS_RIGHT},
                            "installateur": "ERREUR",
                            "is_avoir": False,
                            "source": error or "Echec extraction",
                        })
                except Exception as e:
                    batch_results4.append({
                        "filename": short_name,
                        "fields": {k: None for k, _ in FIELDS_LEFT + FIELDS_RIGHT},
                        "installateur": "ERREUR",
                        "is_avoir": False,
                        "source": str(e),
                    })

            progress4.progress(1.0, text="Termine !")
            progress4.empty()

            # Statistiques
            nb_ok4 = sum(1 for r in batch_results4 if r["installateur"] != "ERREUR")
            nb_err4 = len(batch_results4) - nb_ok4
            installateurs4 = set(r["installateur"] for r in batch_results4 if r["installateur"] != "ERREUR")
            nb_missing4 = sum(
                1 for r in batch_results4 if r["installateur"] != "ERREUR"
                for k, _ in FIELDS_LEFT + FIELDS_RIGHT
                if not r["fields"].get(k) or r["fields"].get(k) == "null"
            )

            st.markdown(
                f'<div class="stats-row">'
                f'<div class="stat-box"><div class="num">{nb_ok4}/{len(batch_results4)}</div>'
                f'<div class="lbl">Factures traitees</div></div>'
                f'<div class="stat-box"><div class="num">{len(installateurs4)}</div>'
                f'<div class="lbl">Installateurs</div></div>'
                f'<div class="stat-box"><div class="num">{nb_missing4}</div>'
                f'<div class="lbl">Champs manquants</div></div>'
                f'<div class="stat-box"><div class="num">{nb_err4}</div>'
                f'<div class="lbl">Erreurs</div></div>'
                f'</div>',
                unsafe_allow_html=True,
            )

            # Tableau recapitulatif
            st.markdown('<div class="section-title">Resultats par facture</div>', unsafe_allow_html=True)

            for res in batch_results4:
                is_err = res["installateur"] == "ERREUR"
                icon = "❌" if is_err else "✅"
                src = res["source"]
                with st.expander(
                    f"{icon} {res['filename']} — {res['installateur'].upper()} ({src})",
                    expanded=False,
                ):
                    if is_err:
                        st.error(f"Extraction echouee : {src}")
                    else:
                        show_results(res["fields"])

            # Export Excel multi-feuilles
            st.markdown("---")
            zip_base4 = os.path.splitext(uploaded_zip4.name)[0]
            st.download_button(
                f"Telecharger Excel — {len(batch_results4)} factures ({len(installateurs4)} feuilles)",
                data=export_excel_multi_sheets(batch_results4),
                file_name=f"{zip_base4}_batch_multi.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel_tab4",
            )


# ─────────────────────────────────────────────────────────────
# FOOTER
# ─────────────────────────────────────────────────────────────
st.markdown(
    '<div class="app-footer">Stage 2026 — MADANI Yassine · Gemma2:9b</div>',
    unsafe_allow_html=True,
)
